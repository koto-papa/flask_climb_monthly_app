from flask import render_template, url_for, redirect, flash, request, abort, send_file
from result import db
from result.forms import RegistrationForm, LoginForm, UpdateUserForm
from result.models import User
from flask import Blueprint
from flask_login import login_user, logout_user, login_required, current_user
import openpyxl as op
from io import BytesIO


users = Blueprint('users', __name__)

@users.route('/')
def index():
    return render_template('index.html')

@users.route('/register', methods=['GET','POST'])
def register():
    form = RegistrationForm()
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            password=form.password.data,
            result=form.result.data,
            administrator='0'
            )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash('登録が完了しました。')
        return redirect(url_for('users.register'))
    return render_template('register.html',form=form)

@users.route('/add_admin', methods=['GET', 'POST'])
def add_admin():
    if request.method == 'GET':
        return render_template('add_admin.html')
    if request.method == 'POST':
        db.create_all()
        admin = User(
            username='admin',
            password='123',
            result='1',
            administrator='1'
            )
        admin.set_password('123')
        try:
            db.session.add(admin)
            db.session.commit()
            flash('登録が完了しました。')
            return redirect(url_for('users.add_admin'))
        except Exception:
            flash('すでに登録されています。')
            return redirect(url_for('users.add_admin'))

@users.route('/result_all')
@login_required
def result_all():
    users = User.query.order_by(User.result.desc()).all()
    return render_template('result_all.html', users=users)

@users.route('/login', methods=['GET','POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is not None:
            if user.check_password(form.password.data):
                login_user(user)
                next = request.args.get('next')
                if next == None or not next[0] == '/':
                    next = url_for('users.result_all')
                return redirect(next)
            else:
                flash('パスワードが一致しません。')
        else:
            flash('入力されたユーザーは存在しません。')
    return render_template('login.html', form=form)

@users.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('users.index'))

@users.route('/<int:user_id>/account', methods=['GET', 'POST'])
@login_required
def account(user_id):    
    user = User.query.get_or_404(user_id)
    if user.id != current_user.id and not current_user.is_administrator():
        abort(403)
    form = UpdateUserForm(user_id)
    if form.validate_on_submit():
        user.result = form.result.data
        db.session.commit()
        flash('完登数が更新されました。')
        return redirect(url_for('users.result_all'))
    return render_template('account.html', form=form)

@users.route('/<int:user_id>/delete', methods=['GET', 'POST'])
@login_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if not current_user.is_administrator():
        abort(403)
    db.session.delete(user)
    db.session.commit()
    flash('登録を削除しました。')
    return redirect(url_for('users.result_all'))

@users.route('/delete_all', methods=['GET', 'POST'])
@login_required
def delete_all():
    users = User.query.filter(User.administrator==0).all()
    if not current_user.is_administrator():
        abort(403)
    for i in users:
        db.session.delete(i)
    db.session.commit()
    flash('すべての登録を削除しました。')
    return redirect(url_for('users.result_all'))

@users.route('/download')
@login_required
def download():
    if not current_user.is_administrator():
        abort(403)     
    wb = op.Workbook()
    ws = wb.worksheets[0]   
    ws.cell(1, 1).value = 'ニックネーム'
    ws.cell(1, 2).value = '完登数'
    users = User.query.order_by(User.result.desc()).all() 
    for j, user in enumerate(users):       
        ws.cell(j+2, 1).value = user.username 
        ws.cell(j+2, 2).value = user.result
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    name = 'result.xlsx'
    return send_file(output, attachment_filename=name, as_attachment=True)
